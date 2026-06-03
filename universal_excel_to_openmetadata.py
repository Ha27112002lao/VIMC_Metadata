#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Universal Excel -> OpenMetadata Glossary importer.

Goal
----
This script is designed to RUN with almost any Excel workbook, even when the
file is not in the original VIMC glossary template.

It works in 2 levels:
1. Structured glossary/table files:
   - Auto-detects header rows and common columns such as Code, Name, Term,
     Group, Definition, Unit, Source, Formula, Owner, Tags, etc.
   - Creates a clean Glossary hierarchy with ordered group terms and row terms.
2. Unstructured/random Excel files:
   - Still reads every non-empty row from each sheet.
   - Generates safe term names from row order and available cell values.
   - Stores the original row/cell values in the term description.

Important truth
---------------
No program can understand the business meaning of a completely random Excel
file perfectly. This script's promise is:
- it should not require a fixed column order;
- it should preserve workbook/sheet/row order;
- it should not stop just because the template is different;
- it should create reasonable OpenMetadata glossary terms by best effort.

Default OpenMetadata hierarchy
------------------------------
<Glossary>
└── W0001_<Workbook name>
    └── S0001_<Sheet name>
        ├── G0001_<Group name>        # only when group/category is detected
        │   ├── T000001_<Term/code>
        │   └── T000002_<Term/code>
        └── G0002_Ungrouped
            └── T000003_<Term/code>

The W/S/G/T prefixes preserve the original Excel order in UIs that sort terms by
name. The displayName still keeps the human-readable workbook/sheet/group/term.

Environment variables
---------------------
OPENMETADATA_HOST   Example: http://192.168.74.12:30085
OM_JWT_TOKEN        Personal Access Token or Bot Token

Examples
--------
# Check what will be imported without calling OpenMetadata
python universal_excel_to_openmetadata.py --dry-run --excel "any_file.xlsx"

# Import one file
python universal_excel_to_openmetadata.py --excel "any_file.xlsx"

# Import many files
python universal_excel_to_openmetadata.py --excel file1.xlsx --excel file2.xlsx

# Keep old style options from previous scripts
python universal_excel_to_openmetadata.py --excel "Business Glossary Demo.xlsx" --glossary VIMC_Business_Glossary

# Import only one sheet
python universal_excel_to_openmetadata.py --excel "any_file.xlsx" --sheet "Logic Tổng hợp"

# Disable order prefixes in OpenMetadata entity names
python universal_excel_to_openmetadata.py --excel "any_file.xlsx" --name-mode original
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import sys
import time
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


DEFAULT_GLOSSARY = "VIMC_Business_Glossary"
DEFAULT_HEADER_SCAN_ROWS = 40
DEFAULT_MAX_COLS = 300
DEFAULT_MAX_EMPTY_ROWS_AFTER_DATA = 30
DEFAULT_MAX_TERMS_PER_SHEET = 10000
DEFAULT_TIMEOUT_SECONDS = 45
DEFAULT_SHEET_CANDIDATES = (
    "Logic Tổng hợp",
    "Logic Tong hop",
    "Glossary",
    "Business Glossary",
    "Terms",
    "Data",
)

EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
CSV_EXTENSIONS = {".csv", ".tsv"}
INVALID_VALUES = {
    "", "-", "--", "---", "n/a", "na", "none", "null", "nil", "nan", "#n/a",
    "#value!", "#ref!", "#div/0!", "#name?", "#num!", "#null!", "#spill!",
}


@dataclass(frozen=True)
class FieldSpec:
    key: str
    label: str
    aliases: Tuple[str, ...]
    weight: int = 1


FIELD_SPECS: Tuple[FieldSpec, ...] = (
    FieldSpec("order", "STT/Order", ("stt", "no", "no.", "so thu tu", "thu tu", "order", "ordinal", "index", "seq", "sequence"), 1),
    FieldSpec("code", "Mã/Code", ("code", "ma", "mã", "ma chi tieu", "mã chỉ tiêu", "ma thuat ngu", "mã thuật ngữ", "term code", "business code", "id", "identifier"), 5),
    FieldSpec("term", "Tên term", ("term", "term name", "glossary term", "business term", "name", "ten", "tên", "ten term", "tên term", "ten thuat ngu", "tên thuật ngữ", "chi tieu", "chỉ tiêu", "indicator", "metric", "kpi", "field", "attribute"), 5),
    FieldSpec("display_name", "Display name", ("display name", "displayname", "display_name", "label", "title", "ten hien thi", "tên hiển thị", "indicator new", "indicator moi", "chi tieu moi", "chỉ tiêu mới", "ten moi", "tên mới"), 4),
    FieldSpec("group", "Nhóm/Parent", ("group", "nhom", "nhóm", "nhom chi tieu", "nhóm chỉ tiêu", "category", "parent", "parent term", "parent group", "domain", "subject", "topic", "chu de", "chủ đề", "nhom du lieu", "nhóm dữ liệu", "section"), 6),
    FieldSpec("indicator_scope", "Indicator scope", ("indicator file scope", "indicator scope", "file scope", "scope", "pham vi", "phạm vi"), 4),
    FieldSpec("definition", "Định nghĩa/Description", ("definition", "description", "meaning", "business definition", "mo ta", "mô tả", "dinh nghia", "định nghĩa", "dien giai", "diễn giải", "noi dung", "nội dung"), 5),
    FieldSpec("unit", "Đơn vị tính", ("unit", "uom", "unit of measure", "don vi", "đơn vị", "don vi tinh", "đơn vị tính"), 2),
    FieldSpec("source", "Nguồn", ("source", "source name", "ten nguon", "tên nguồn", "nguon", "nguồn", "system", "application", "app", "database", "table", "raw table", "silver table", "gold table"), 2),
    FieldSpec("formula", "Công thức", ("formula", "calculation", "calc", "logic", "business logic", "cong thuc", "công thức", "cach tinh", "cách tính", "công thức tính toán", "formula logic"), 2),
    FieldSpec("owner", "Owner", ("owner", "data owner", "steward", "data steward", "business owner", "pic", "phu trach", "phụ trách", "nguoi phu trach", "người phụ trách"), 2),
    FieldSpec("note", "Ghi chú", ("note", "notes", "comment", "comments", "ghi chu", "ghi chú", "remark", "remarks"), 1),
    FieldSpec("priority", "Priority", ("priority", "muc do uu tien", "mức độ ưu tiên", "uu tien", "ưu tiên"), 1),
    FieldSpec("tag", "Tags", ("tag", "tags", "classification", "classifications", "phan loai", "phân loại"), 1),
)

FIELD_BY_KEY = {f.key: f for f in FIELD_SPECS}


# ---------------------------------------------------------------------------
# Generic helpers
# ---------------------------------------------------------------------------

def clean(value: Any) -> str:
    """Convert cell values to safe, readable strings."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return (f"{value:.12g}").strip()
    text = str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    return text.strip()


def is_meaningful(value: Any) -> bool:
    text = clean(value)
    if not text:
        return False
    return text.strip().lower() not in INVALID_VALUES


def strip_accents(text: str) -> str:
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return text.replace("đ", "d").replace("Đ", "D")


def normalize_token(text: Any) -> str:
    text = strip_accents(clean(text)).lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def safe_name(text: Any, fallback: str = "Term", max_len: int = 96) -> str:
    """Create an OpenMetadata-friendly entity name."""
    raw = clean(text)
    if not raw:
        raw = fallback
    text = strip_accents(raw)
    text = re.sub(r"[^A-Za-z0-9_\-]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_-")
    if not text:
        text = fallback
    if re.match(r"^[0-9]", text):
        text = "N_" + text
    return text[:max_len].strip("_-") or fallback


def short_hash(text: str, length: int = 8) -> str:
    return hashlib.sha1(text.encode("utf-8", errors="ignore")).hexdigest()[:length]


def unique_name(base: str, used: set[str], max_len: int = 110) -> str:
    name = base[:max_len]
    if name not in used:
        used.add(name)
        return name
    digest = short_hash(base)
    for i in range(2, 100000):
        suffix = f"_{i}_{digest}"
        candidate = f"{base[:max_len-len(suffix)]}{suffix}"
        if candidate not in used:
            used.add(candidate)
            return candidate
    candidate = f"{base[:max_len-13]}_{digest}_{len(used)}"
    used.add(candidate)
    return candidate


def fqn_join(*parts: str) -> str:
    return ".".join(p for p in parts if p)


def truncate(text: str, max_len: int = 36000) -> str:
    if len(text) <= max_len:
        return text
    return text[: max_len - 80] + "\n\n...[truncated because the row is too large]"


def make_ordered_name(prefix: str, order: int, raw: Any, fallback: str, name_mode: str) -> str:
    base = safe_name(raw, fallback=fallback)
    if name_mode == "original":
        return base
    width = 6 if prefix == "T" else 4
    return f"{prefix}{order:0{width}d}_{base}"


def split_multi_value(text: str) -> List[str]:
    parts = re.split(r"[,;|\n]+", clean(text))
    return [p.strip() for p in parts if is_meaningful(p)]


# ---------------------------------------------------------------------------
# Reading workbooks without requiring one fixed template
# ---------------------------------------------------------------------------

@dataclass
class SheetData:
    source_file: Path
    sheet_name: str
    rows: List[List[str]]


@dataclass
class DetectedTable:
    header_row_index: Optional[int]       # 0-based index in SheetData.rows, None when generated
    data_start_index: int                 # 0-based index in SheetData.rows
    headers: List[str]
    field_to_col: Dict[str, int]
    score: int
    mode: str                             # structured | generic


def read_excel_workbook(path: Path, sheet_filter: Optional[str], all_sheets: bool) -> List[SheetData]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_filter:
            if sheet_filter not in wb.sheetnames:
                available = ", ".join(wb.sheetnames)
                raise ValueError(f"File '{path.name}' không có sheet '{sheet_filter}'. Sheets hiện có: {available}")
            sheet_names = [sheet_filter]
        elif all_sheets:
            sheet_names = list(wb.sheetnames)
        else:
            sheet_names = choose_default_sheet_names(wb.sheetnames)

        results: List[SheetData] = []
        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            rows: List[List[str]] = []
            # Some Excel files have stale worksheet dimensions in read_only mode.
            # Reading up to DEFAULT_MAX_COLS is safer than trusting ws.max_column.
            max_col_seen = DEFAULT_MAX_COLS
            for raw_row in ws.iter_rows(max_col=max_col_seen, values_only=True):
                row = [clean(v) for v in raw_row]
                # trim only trailing empty cells; keep internal blanks for column positions
                while row and not is_meaningful(row[-1]):
                    row.pop()
                rows.append(row)
            results.append(SheetData(path, sheet_name, rows))
        return results
    finally:
        wb.close()


def choose_default_sheet_names(sheet_names: Sequence[str]) -> List[str]:
    normalized = {normalize_token(s): s for s in sheet_names}
    for candidate in DEFAULT_SHEET_CANDIDATES:
        n = normalize_token(candidate)
        if n in normalized:
            return [normalized[n]]
    return [sheet_names[0]] if sheet_names else []


def read_csv_file(path: Path) -> List[SheetData]:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    rows: List[List[str]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, delimiter=delimiter)
        for row in reader:
            cleaned = [clean(v) for v in row]
            while cleaned and not is_meaningful(cleaned[-1]):
                cleaned.pop()
            rows.append(cleaned)
    return [SheetData(path, path.stem, rows)]


def read_input_file(path: Path, sheet_filter: Optional[str], all_sheets: bool) -> List[SheetData]:
    suffix = path.suffix.lower()
    if suffix in EXCEL_EXTENSIONS:
        return read_excel_workbook(path, sheet_filter=sheet_filter, all_sheets=all_sheets)
    if suffix in CSV_EXTENSIONS:
        return read_csv_file(path)
    if suffix == ".xls":
        raise ValueError(
            f"File '{path.name}' là định dạng .xls cũ. Hãy Save As sang .xlsx rồi chạy lại. "
            "Script này dùng openpyxl nên hỗ trợ .xlsx/.xlsm/.xltx/.xltm."
        )
    raise ValueError(f"Không hỗ trợ file '{path.name}'. Hỗ trợ: .xlsx, .xlsm, .xltx, .xltm, .csv, .tsv")


def row_non_empty_count(row: Sequence[str]) -> int:
    return sum(1 for v in row if is_meaningful(v))


def value_at(row: Sequence[str], idx: Optional[int]) -> str:
    if idx is None or idx < 0 or idx >= len(row):
        return ""
    return clean(row[idx])


def field_match_score(header_value: str, spec: FieldSpec) -> int:
    token = normalize_token(header_value)
    if not token:
        return 0
    best = 0
    for alias in spec.aliases:
        a = normalize_token(alias)
        if not a:
            continue
        # Exact header match is strongest.
        if token == a:
            best = max(best, 100 * spec.weight + len(a))
            continue
        # Phrase match with word boundaries. Longer aliases are more specific.
        if re.search(r"(^| )" + re.escape(a) + r"( |$)", token):
            best = max(best, 45 * spec.weight + len(a))
            continue
        # Weak fallback for abbreviated headers.
        if len(token) >= 3 and token in a:
            best = max(best, 20 * spec.weight + len(token))
    return best


def detect_field_mapping(headers: Sequence[str]) -> Tuple[Dict[str, int], int]:
    """Map known fields to columns using aliases."""
    field_to_col: Dict[str, int] = {}
    total_score = 0
    used_cols: set[int] = set()
    candidates: List[Tuple[int, str, int]] = []
    for col_idx, header in enumerate(headers):
        for spec in FIELD_SPECS:
            score = field_match_score(header, spec)
            if score:
                candidates.append((score, spec.key, col_idx))
    # Highest score first; for ties prefer the earlier column because
    # glossary files usually put the primary term/code columns before metadata columns.
    candidates.sort(key=lambda x: (-x[0], x[2]))
    for score, key, col_idx in candidates:
        if key in field_to_col or col_idx in used_cols:
            continue
        field_to_col[key] = col_idx
        used_cols.add(col_idx)
        total_score += score
    return field_to_col, total_score


def build_combined_headers(rows: Sequence[Sequence[str]], header_idx: int) -> List[str]:
    """Combine the chosen header row with the previous row when both look like header rows.

    This handles files where the header is split across two rows. For example,
    row 3 may contain English labels and row 4 may contain Vietnamese labels.
    """
    current = list(rows[header_idx])
    previous = list(rows[header_idx - 1]) if header_idx > 0 else []
    max_len = max(len(current), len(previous), 1)
    headers: List[str] = []
    for i in range(max_len):
        cur = current[i] if i < len(current) else ""
        prev = previous[i] if i < len(previous) else ""
        if is_meaningful(prev) and is_meaningful(cur) and normalize_token(prev) != normalize_token(cur):
            prev_count = row_non_empty_count(previous)
            if prev_count >= 3:
                headers.append(f"{prev} / {cur}")
            else:
                headers.append(cur)
        else:
            headers.append(cur or prev or f"Column {get_column_letter(i + 1)}")
    return headers


def combine_header_rows(top: Sequence[str], bottom: Sequence[str]) -> List[str]:
    """Combine two adjacent header rows, top first then bottom."""
    max_len = max(len(top), len(bottom), 1)
    headers: List[str] = []
    for i in range(max_len):
        a = top[i] if i < len(top) else ""
        b = bottom[i] if i < len(bottom) else ""
        if is_meaningful(a) and is_meaningful(b) and normalize_token(a) != normalize_token(b):
            headers.append(f"{a} / {b}")
        else:
            headers.append(a or b or f"Column {get_column_letter(i + 1)}")
    return headers


def row_alias_score(row: Sequence[str]) -> int:
    mapping, score = detect_field_mapping([clean(v) for v in row])
    return score


def maybe_extend_header_down(rows: Sequence[Sequence[str]], table: DetectedTable) -> DetectedTable:
    """If the row after the detected header is another header row, merge it.

    This prevents bilingual two-row headers from being imported as data.
    """
    if table.header_row_index is None:
        return table
    next_idx = table.header_row_index + 1
    if next_idx >= len(rows):
        return table
    next_row = rows[next_idx]
    if row_non_empty_count(next_row) < 2:
        return table
    next_score = row_alias_score(next_row)
    # Treat the next row as a header only if it contains recognizable header aliases.
    if next_score < 100:
        return table
    combined = combine_header_rows(rows[table.header_row_index], next_row)
    mapping, score = detect_field_mapping(combined)
    if score >= table.score or next_score >= 100:
        return DetectedTable(
            header_row_index=table.header_row_index,
            data_start_index=next_idx + 1,
            headers=combined,
            field_to_col=mapping,
            score=max(score, table.score),
            mode="structured" if score >= 100 else table.mode,
        )
    return table


def detect_table(sheet: SheetData, forced_header_row: Optional[int], forced_data_start_row: Optional[int]) -> DetectedTable:
    rows = sheet.rows
    if not rows:
        return DetectedTable(None, 0, [], {}, 0, "generic")

    if forced_header_row is not None:
        idx = max(0, forced_header_row - 1)
        headers = build_combined_headers(rows, idx)
        mapping, score = detect_field_mapping(headers)
        data_start = forced_data_start_row - 1 if forced_data_start_row else idx + 1
        return maybe_extend_header_down(rows, DetectedTable(idx, max(data_start, 0), headers, mapping, score, "structured" if score else "generic"))

    best: Optional[DetectedTable] = None
    scan_limit = min(len(rows), DEFAULT_HEADER_SCAN_ROWS)
    for idx in range(scan_limit):
        row = rows[idx]
        non_empty = row_non_empty_count(row)
        if non_empty < 2:
            continue
        headers = build_combined_headers(rows, idx)
        mapping, alias_score = detect_field_mapping(headers)
        # Give a small bonus to rows with multiple non-empty cells and text-like headers.
        unique_headers = len({normalize_token(h) for h in headers if normalize_token(h)})
        structural_score = min(non_empty, 20) * 5 + min(unique_headers, 20) * 3
        # Penalize rows that look like data values only.
        data_like_penalty = 0
        for h in row:
            n = normalize_token(h)
            if re.fullmatch(r"\d+(\.\d+)?", n or ""):
                data_like_penalty += 8
        score = alias_score + structural_score - data_like_penalty
        candidate = DetectedTable(idx, idx + 1, headers, mapping, score, "structured" if alias_score >= 100 else "generic")
        if best is None or candidate.score > best.score:
            best = candidate

    if best and best.score >= 35:
        return maybe_extend_header_down(rows, best)

    # No useful header found. Generate generic headers from max row width.
    max_width = max((len(r) for r in rows), default=0)
    headers = [f"Column {get_column_letter(i + 1)}" for i in range(max_width)]
    return DetectedTable(None, 0, headers, {}, 0, "generic")


def is_importable_row(row: Sequence[str], table: DetectedTable) -> bool:
    """Return False for blank/formula-error/header-like rows.

    For structured sheets, a row should have at least one meaningful identity field
    such as code/name/definition. This skips rows like: STT + #N/A + blank cells.
    For generic sheets, any meaningful row is allowed.
    """
    if row_non_empty_count(row) == 0:
        return False

    identity_fields = ["code", "term", "display_name", "definition"]
    identity_indices = [table.field_to_col.get(k) for k in identity_fields if k in table.field_to_col]
    if identity_indices:
        if any(is_meaningful(value_at(row, idx)) for idx in identity_indices):
            return True
        order_idx = table.field_to_col.get("order")
        other_meaningful = [v for i, v in enumerate(row) if i != order_idx and is_meaningful(v)]
        return len(other_meaningful) >= 2

    return row_non_empty_count(row) >= 1


def iter_data_rows(sheet: SheetData, table: DetectedTable, max_terms: int) -> Iterable[Tuple[int, List[str]]]:
    """Yield (1-based Excel row number, row values)."""
    empty_streak = 0
    emitted = 0
    for idx in range(table.data_start_index, len(sheet.rows)):
        row = list(sheet.rows[idx])
        if not is_importable_row(row, table):
            empty_streak += 1
            if empty_streak >= DEFAULT_MAX_EMPTY_ROWS_AFTER_DATA and emitted > 0:
                break
            continue
        empty_streak = 0
        emitted += 1
        yield idx + 1, row
        if emitted >= max_terms:
            print(f"⚠️ Sheet '{sheet.sheet_name}' đạt giới hạn {max_terms} terms, bỏ qua phần còn lại.")
            break


# ---------------------------------------------------------------------------
# Building OpenMetadata terms
# ---------------------------------------------------------------------------

@dataclass
class TermPlan:
    name: str
    display_name: str
    description: str
    parent_fqn: Optional[str]
    synonyms: List[str]


@dataclass
class SheetPlan:
    workbook_term: TermPlan
    sheet_term: TermPlan
    group_terms: List[TermPlan]
    row_terms: List[TermPlan]
    detected_table: DetectedTable


def row_to_description(
    source_file: Path,
    sheet_name: str,
    excel_row_number: int,
    headers: Sequence[str],
    row: Sequence[str],
    field_to_col: Dict[str, int],
) -> str:
    lines = [
        f"**Source file:** {source_file.name}",
        f"**Sheet:** {sheet_name}",
        f"**Excel row:** {excel_row_number}",
    ]

    # Highlight common business fields first.
    for key in ("code", "term", "display_name", "group", "definition", "unit", "source", "formula", "owner", "priority", "note", "tag"):
        col_idx = field_to_col.get(key)
        val = value_at(row, col_idx)
        if is_meaningful(val):
            lines.append(f"**{FIELD_BY_KEY[key].label}:** {val}")

    # Then include the full original row, preserving column order.
    details: List[str] = []
    max_len = max(len(headers), len(row))
    for i in range(max_len):
        value = value_at(row, i)
        if not is_meaningful(value):
            continue
        header = headers[i] if i < len(headers) and is_meaningful(headers[i]) else f"Column {get_column_letter(i + 1)}"
        details.append(f"- **{header}:** {value}")
    if details:
        lines.append("\n**Original row values:**")
        lines.extend(details)
    return truncate("\n\n".join(lines))


def choose_row_identity(row: Sequence[str], headers: Sequence[str], mapping: Dict[str, int], excel_row_number: int) -> Tuple[str, str, List[str]]:
    """Return raw_name, display_name, synonyms for a row term."""
    code = value_at(row, mapping.get("code"))
    term = value_at(row, mapping.get("term"))
    display = value_at(row, mapping.get("display_name"))
    definition = value_at(row, mapping.get("definition"))

    # Best effort for files without known headers: use the first meaningful cell,
    # then the second one as a more descriptive display if available.
    meaningful_values = [clean(v) for v in row if is_meaningful(v)]
    fallback = meaningful_values[0] if meaningful_values else f"row_{excel_row_number}"

    raw_name = code or term or display or fallback or f"row_{excel_row_number}"
    display_name = display or term or code or fallback or f"Row {excel_row_number}"
    if not is_meaningful(display_name) and definition:
        display_name = definition[:120]

    synonyms: List[str] = []
    for candidate in (code, term, display):
        if is_meaningful(candidate) and candidate != display_name:
            synonyms.append(candidate)
    # De-duplicate while preserving order.
    deduped: List[str] = []
    seen: set[str] = set()
    for s in synonyms:
        if s not in seen:
            deduped.append(s)
            seen.add(s)
    return raw_name, display_name, deduped[:10]


def choose_group(row: Sequence[str], mapping: Dict[str, int]) -> str:
    group = value_at(row, mapping.get("group"))
    if is_meaningful(group):
        return group
    return "Ungrouped"


def build_sheet_plan(
    sheet: SheetData,
    table: DetectedTable,
    glossary_name: str,
    workbook_order: int,
    sheet_order: int,
    name_mode: str,
    max_terms_per_sheet: int,
) -> SheetPlan:
    workbook_raw = sheet.source_file.stem
    workbook_name = make_ordered_name("W", workbook_order, workbook_raw, "Workbook", name_mode)
    workbook_fqn = fqn_join(glossary_name, workbook_name)
    workbook_term = TermPlan(
        name=workbook_name,
        display_name=workbook_raw,
        description=f"Imported from workbook: {sheet.source_file.name}",
        parent_fqn=None,
        synonyms=[],
    )

    sheet_raw = sheet.sheet_name
    sheet_name = make_ordered_name("S", sheet_order, sheet_raw, "Sheet", name_mode)
    sheet_fqn = fqn_join(glossary_name, workbook_name, sheet_name)
    sheet_term = TermPlan(
        name=sheet_name,
        display_name=sheet_raw,
        description=(
            f"Sheet '{sheet.sheet_name}' imported from workbook '{sheet.source_file.name}'.\n\n"
            f"Detected mode: {table.mode}.\n"
            f"Header row: {table.header_row_index + 1 if table.header_row_index is not None else 'generated/no header'}.\n"
            f"Data starts at row: {table.data_start_index + 1}."
        ),
        parent_fqn=workbook_fqn,
        synonyms=[],
    )

    used_group_names: set[str] = set()
    used_row_names: set[str] = set()
    group_terms_by_key: Dict[str, TermPlan] = {}
    row_terms: List[TermPlan] = []
    group_order_by_key: Dict[str, int] = {}

    for row_order, (excel_row_number, row) in enumerate(iter_data_rows(sheet, table, max_terms_per_sheet), start=1):
        group_raw = choose_group(row, table.field_to_col)
        group_key = normalize_token(group_raw) or "ungrouped"
        if group_key not in group_order_by_key:
            group_order_by_key[group_key] = len(group_order_by_key) + 1
            group_name_base = make_ordered_name("G", group_order_by_key[group_key], group_raw, "Group", name_mode)
            group_name = unique_name(group_name_base, used_group_names)
            group_fqn = fqn_join(glossary_name, workbook_name, sheet_name, group_name)
            group_terms_by_key[group_key] = TermPlan(
                name=group_name,
                display_name=group_raw,
                description=f"Group/category '{group_raw}' in sheet '{sheet.sheet_name}' from workbook '{sheet.source_file.name}'.",
                parent_fqn=sheet_fqn,
                synonyms=[],
            )
        else:
            group_name = group_terms_by_key[group_key].name
            group_fqn = fqn_join(glossary_name, workbook_name, sheet_name, group_name)

        raw_name, display_name, synonyms = choose_row_identity(row, table.headers, table.field_to_col, excel_row_number)
        term_name_base = make_ordered_name("T", row_order, raw_name, f"row_{excel_row_number}", name_mode)
        term_name = unique_name(term_name_base, used_row_names)
        description = row_to_description(
            source_file=sheet.source_file,
            sheet_name=sheet.sheet_name,
            excel_row_number=excel_row_number,
            headers=table.headers,
            row=row,
            field_to_col=table.field_to_col,
        )
        row_terms.append(
            TermPlan(
                name=term_name,
                display_name=display_name,
                description=description,
                parent_fqn=group_fqn,
                synonyms=synonyms,
            )
        )

    return SheetPlan(
        workbook_term=workbook_term,
        sheet_term=sheet_term,
        group_terms=list(group_terms_by_key.values()),
        row_terms=row_terms,
        detected_table=table,
    )


# ---------------------------------------------------------------------------
# OpenMetadata API client
# ---------------------------------------------------------------------------

class OpenMetadataClient:
    def __init__(self, host: str, token: Optional[str], timeout: int = DEFAULT_TIMEOUT_SECONDS):
        self.base_url = self.normalize_api_base(host)
        self.timeout = timeout
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
        if token:
            self.session.headers.update({"Authorization": f"Bearer {token}"})

    @staticmethod
    def normalize_api_base(host: str) -> str:
        if not host:
            raise ValueError("OPENMETADATA_HOST is empty. Example: http://192.168.74.12:30085")
        host = host.strip().rstrip("/")
        if host.endswith("/api/v1"):
            return host
        if host.endswith("/api"):
            return host + "/v1"
        return host + "/api/v1"

    def request(self, method: str, path: str, json_body: Optional[dict] = None) -> dict:
        url = f"{self.base_url}{path}"
        response = self.session.request(method, url, json=json_body, timeout=self.timeout)
        if response.status_code >= 400:
            hint = ""
            if response.status_code == 401:
                hint = "\nGợi ý: token thiếu/sai. Kiểm tra OM_JWT_TOKEN."
            elif response.status_code == 403:
                hint = "\nGợi ý: token không đủ quyền tạo/sửa Glossary."
            elif response.status_code == 400:
                hint = "\nGợi ý: payload có thể chưa khớp version OpenMetadata. Xem response body."
            elif response.status_code == 404:
                hint = "\nGợi ý: kiểm tra OPENMETADATA_HOST có đúng host OpenMetadata và có /api/v1 không."
            raise RuntimeError(
                f"\nAPI ERROR {response.status_code} {method} {url}\n"
                f"Request body: {json.dumps(json_body, ensure_ascii=False) if json_body else ''}\n"
                f"Response body: {response.text}{hint}\n"
            )
        if not response.text.strip():
            return {}
        try:
            return response.json()
        except Exception:
            return {"raw": response.text}

    def get(self, path: str) -> dict:
        return self.request("GET", path)

    def put(self, path: str, body: dict) -> dict:
        return self.request("PUT", path, body)

    def health_check(self) -> None:
        try:
            version = self.get("/system/version")
            print(f"✅ Connected to OpenMetadata: {self.base_url}")
            if version:
                print(f"   Version: {version}")
        except Exception as exc:
            print("⚠️ Không gọi được /system/version. Script vẫn sẽ thử import tiếp.")
            print(f"   Chi tiết: {exc}")

    def upsert_glossary(self, name: str, display_name: str, description: str, dry_run: bool) -> None:
        payload = {
            "name": name,
            "displayName": display_name,
            "description": description,
            "mutuallyExclusive": False,
        }
        if dry_run:
            print(f"[DRY-RUN] Upsert glossary: {name}")
            return
        self.put("/glossaries", payload)
        print(f"✅ Upsert glossary: {name}")

    def upsert_term(self, glossary: str, term: TermPlan, dry_run: bool) -> None:
        payload: Dict[str, Any] = {
            "name": term.name,
            "glossary": glossary,
            "displayName": term.display_name or term.name,
            "description": term.description or f"Imported term: {term.display_name or term.name}",
        }
        if term.parent_fqn:
            payload["parent"] = term.parent_fqn
        if term.synonyms:
            payload["synonyms"] = term.synonyms

        if dry_run:
            parent_note = f" under {term.parent_fqn}" if term.parent_fqn else ""
            print(f"[DRY-RUN] Upsert term: {glossary}.{term.name}{parent_note}")
            return
        self.put("/glossaryTerms", payload)
        parent_note = f" under {term.parent_fqn}" if term.parent_fqn else ""
        print(f"✅ Upsert term: {glossary}.{term.name}{parent_note}")


# ---------------------------------------------------------------------------
# Import execution
# ---------------------------------------------------------------------------

def resolve_input_paths(paths: Iterable[str], auto_scan: bool) -> List[Path]:
    resolved: List[Path] = []
    for p in paths:
        path = Path(p)
        if not path.exists():
            raise FileNotFoundError(f"Không tìm thấy file: {p}")
        if path.is_dir():
            for ext in sorted(EXCEL_EXTENSIONS | CSV_EXTENSIONS):
                resolved.extend(sorted(path.glob(f"*{ext}")))
        else:
            resolved.append(path)

    if not resolved and auto_scan:
        cwd = Path.cwd()
        for ext in sorted(EXCEL_EXTENSIONS | CSV_EXTENSIONS):
            resolved.extend(sorted(cwd.glob(f"*{ext}")))

    # de-duplicate while preserving order
    deduped: List[Path] = []
    seen: set[Path] = set()
    for p in resolved:
        rp = p.resolve()
        if rp not in seen:
            deduped.append(p)
            seen.add(rp)
    if not deduped:
        raise FileNotFoundError("Không có file Excel/CSV đầu vào. Hãy dùng --excel 'file.xlsx'.")
    return deduped


def print_detection_summary(sheet: SheetData, table: DetectedTable) -> None:
    mapped = ", ".join(f"{k}=col{v+1}" for k, v in sorted(table.field_to_col.items(), key=lambda x: x[1])) or "none"
    header = table.header_row_index + 1 if table.header_row_index is not None else "generated"
    print(
        f"   • Sheet '{sheet.sheet_name}': mode={table.mode}, "
        f"header={header}, data_start={table.data_start_index + 1}, mapped_fields=[{mapped}]"
    )


def execute_import(
    client: OpenMetadataClient,
    plans: Sequence[SheetPlan],
    glossary_name: str,
    dry_run: bool,
    sleep_seconds: float,
) -> Tuple[int, int, int]:
    created_workbooks: set[str] = set()
    workbook_count = 0
    sheet_count = 0
    group_count = 0
    row_count = 0

    for plan in plans:
        if plan.workbook_term.name not in created_workbooks:
            client.upsert_term(glossary_name, plan.workbook_term, dry_run=dry_run)
            created_workbooks.add(plan.workbook_term.name)
            workbook_count += 1
            if sleep_seconds:
                time.sleep(sleep_seconds)

        client.upsert_term(glossary_name, plan.sheet_term, dry_run=dry_run)
        sheet_count += 1
        if sleep_seconds:
            time.sleep(sleep_seconds)

        for group_term in plan.group_terms:
            client.upsert_term(glossary_name, group_term, dry_run=dry_run)
            group_count += 1
            if sleep_seconds:
                time.sleep(sleep_seconds)

        for row_term in plan.row_terms:
            client.upsert_term(glossary_name, row_term, dry_run=dry_run)
            row_count += 1
            if sleep_seconds:
                time.sleep(sleep_seconds)

    return sheet_count, group_count, row_count


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Universal Excel/CSV to OpenMetadata Glossary importer. Best-effort import for almost any workbook."
    )
    parser.add_argument("files", nargs="*", help="Input Excel/CSV files or directories. Positional style is supported.")
    parser.add_argument("--excel", action="append", default=[], help="Excel/CSV file path. Can be used many times.")
    parser.add_argument("--glossary", default=DEFAULT_GLOSSARY, help=f"OpenMetadata Glossary name. Default: {DEFAULT_GLOSSARY}")
    parser.add_argument("--sheet", default=None, help="Import only this sheet name. Default: import all sheets.")
    parser.add_argument("--first-sheet-only", action="store_true", help="Import only the default/first sheet instead of all sheets.")
    parser.add_argument("--host", default=os.getenv("OPENMETADATA_HOST"), help="OpenMetadata host, e.g. http://192.168.74.12:30085")
    parser.add_argument("--token", default=os.getenv("OM_JWT_TOKEN"), help="OpenMetadata JWT/PAT/Bot token.")
    parser.add_argument("--dry-run", action="store_true", help="Only print what would be imported; do not call OpenMetadata API.")
    parser.add_argument("--sleep", type=float, default=0.0, help="Sleep seconds between API calls, e.g. 0.05 for weak servers.")
    parser.add_argument("--name-mode", choices=("ordered", "original"), default="ordered", help="ordered keeps Excel order with W/S/G/T prefixes. original keeps raw names where possible.")
    parser.add_argument("--header-row", type=int, default=None, help="Force 1-based header row number if auto-detection is wrong.")
    parser.add_argument("--data-start-row", type=int, default=None, help="Force 1-based data start row number.")
    parser.add_argument("--max-terms-per-sheet", type=int, default=DEFAULT_MAX_TERMS_PER_SHEET, help=f"Safety limit per sheet. Default: {DEFAULT_MAX_TERMS_PER_SHEET}.")
    parser.add_argument("--auto-scan-current-folder", action="store_true", help="If no file is passed, scan current folder for Excel/CSV files.")
    args = parser.parse_args()

    input_args = list(args.excel or []) + list(args.files or [])
    try:
        input_paths = resolve_input_paths(input_args, auto_scan=args.auto_scan_current_folder)
    except Exception as exc:
        print(f"❌ {exc}")
        print("Ví dụ: python universal_excel_to_openmetadata.py --dry-run --excel 'file.xlsx'")
        return 1

    glossary_name = safe_name(args.glossary, fallback=DEFAULT_GLOSSARY)
    all_sheets = not args.first_sheet_only

    print("\n=== Universal Excel/CSV → OpenMetadata Glossary Import ===")
    print(f"Glossary: {glossary_name}")
    print(f"Dry-run: {args.dry_run}")
    print(f"Name mode: {args.name_mode}")
    print("Files:")
    for p in input_paths:
        print(f"- {p}")

    if not args.host and not args.dry_run:
        print("\n❌ Thiếu OPENMETADATA_HOST. Ví dụ PowerShell:")
        print("$env:OPENMETADATA_HOST='http://192.168.74.12:30085'")
        return 1
    if not args.token and not args.dry_run:
        print("\n⚠️ OM_JWT_TOKEN đang trống. Nếu server yêu cầu đăng nhập, API có thể lỗi 401.")
        print("PowerShell ví dụ:")
        print("$env:OM_JWT_TOKEN='PASTE_TOKEN_CUA_BAN'")

    client = OpenMetadataClient(args.host or "http://localhost:8585", args.token)
    if not args.dry_run:
        client.health_check()

    client.upsert_glossary(
        name=glossary_name,
        display_name=args.glossary,
        description="Glossary imported automatically from arbitrary Excel/CSV files using universal_excel_to_openmetadata.py.",
        dry_run=args.dry_run,
    )

    plans: List[SheetPlan] = []
    workbook_order_by_path: Dict[Path, int] = {}

    print("\n=== Detecting input structure ===")
    for path in input_paths:
        try:
            sheet_datas = read_input_file(path, sheet_filter=args.sheet, all_sheets=all_sheets)
        except Exception as exc:
            print(f"❌ Bỏ qua file '{path}': {exc}")
            continue
        workbook_order_by_path[path.resolve()] = len(workbook_order_by_path) + 1
        for sheet_order, sheet_data in enumerate(sheet_datas, start=1):
            table = detect_table(sheet_data, args.header_row, args.data_start_row)
            print_detection_summary(sheet_data, table)
            plan = build_sheet_plan(
                sheet=sheet_data,
                table=table,
                glossary_name=glossary_name,
                workbook_order=workbook_order_by_path[path.resolve()],
                sheet_order=sheet_order,
                name_mode=args.name_mode,
                max_terms_per_sheet=args.max_terms_per_sheet,
            )
            if not plan.row_terms:
                print(f"     ⚠️ Sheet '{sheet_data.sheet_name}' không có dòng/cell dữ liệu có nghĩa, bỏ qua term dòng.")
            plans.append(plan)

    if not plans:
        print("\n❌ Không tạo được kế hoạch import nào. Kiểm tra lại file đầu vào.")
        return 1

    print("\n=== Import plan summary ===")
    total_sheets = len(plans)
    total_groups = sum(len(p.group_terms) for p in plans)
    total_rows = sum(len(p.row_terms) for p in plans)
    print(f"Sheets: {total_sheets}")
    print(f"Group terms: {total_groups}")
    print(f"Row/business terms: {total_rows}")

    print("\n=== Executing ===")
    try:
        sheet_count, group_count, row_count = execute_import(
            client=client,
            plans=plans,
            glossary_name=glossary_name,
            dry_run=args.dry_run,
            sleep_seconds=args.sleep,
        )
    except Exception as exc:
        print(f"\n❌ Import lỗi: {exc}")
        return 1

    print("\n🎉 Done")
    print(f"Workbook terms: {len({p.workbook_term.name for p in plans})}")
    print(f"Sheet terms: {sheet_count}")
    print(f"Group terms: {group_count}")
    print(f"Row/business terms: {row_count}")
    if args.dry_run:
        print("Dry-run only: chưa gọi API OpenMetadata.")
    else:
        print(f"Kiểm tra trên OpenMetadata UI: Govern → Glossary → {glossary_name}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
