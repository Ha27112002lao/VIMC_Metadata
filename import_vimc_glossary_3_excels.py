#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Import VIMC/Business Glossary Excel files into OpenMetadata Glossary.

This version keeps the old way of running the script, but removes the hard-coded
Excel column positions. It detects the header row and column order automatically,
so it can import any reasonably structured glossary Excel file.

Default hierarchy on OpenMetadata:

VIMC_Business_Glossary
└── <Excel file name>
    ├── G001_<Group 1>
    │   ├── T000001_<Code or Term name>
    │   └── T000002_<Code or Term name>
    └── G002_<Group 2>
        └── T000010_<Code or Term name>

Important:
- Term names are prefixed by default to preserve the Excel order in UIs that sort
  terms by name. The human-facing displayName remains the original indicator/name.
- Use --name-mode original if you want OpenMetadata entity names to stay as the
  raw Code/Name values without order prefixes.

Environment variables:
- OPENMETADATA_HOST, e.g. http://192.168.74.12:30085
- OM_JWT_TOKEN, Personal Access Token or Bot Token

Examples:
python import_vimc_glossary_3_excels.py --excel "Business Glossary Demo.xlsx"
python import_vimc_glossary_3_excels.py --dry-run --excel "Business Glossary Demo.xlsx"
python import_vimc_glossary_3_excels.py --excel file1.xlsx --excel file2.xlsx --glossary VIMC_Business_Glossary
"""

from __future__ import annotations

import argparse
import os
import re
import sys
import time
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple

import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

DEFAULT_GLOSSARY = "VIMC_Business_Glossary"
DEFAULT_SHEET = "Logic Tổng hợp"
DEFAULT_FILES = ["Business Glossary Demo.xlsx"]
DEFAULT_HEADER_SCAN_ROWS = 30
DEFAULT_MAX_COLS = 250

INVALID_VALUES = {
    "", "-", "--", "---", "n/a", "na", "none", "null", "nan", "#n/a",
    "#value!", "#ref!", "#div/0!", "#name?", "#num!", "#null!",
}


@dataclass(frozen=True)
class FieldSpec:
    key: str
    label: str
    aliases: Tuple[str, ...]
    weight: int = 1


FIELD_SPECS: Tuple[FieldSpec, ...] = (
    FieldSpec("no", "STT", ("no", "stt", "so thu tu", "thu tu", "order", "ordinal", "index"), 1),
    FieldSpec("code", "Mã", ("code", "ma", "ma chi tieu", "ma thuat ngu", "ma term", "term code", "business code", "id"), 4),
    FieldSpec("term_name", "Tên term", ("name", "term", "term name", "glossary term", "business term", "ten term", "ten thuat ngu"), 4),
    FieldSpec("group", "Nhóm/Parent", ("group", "nhom", "nhom chi tieu", "category", "parent", "parent term", "parent group", "domain", "chu de", "nhom du lieu"), 4),
    FieldSpec("display_name", "Display name", ("displayname", "display name", "display_name", "ten hien thi", "label"), 3),
    FieldSpec("business_unit", "Khối áp dụng", ("khoi", "khoi ap dung", "business unit", "unit owner", "department", "division"), 1),
    FieldSpec("indicator_scope", "Indicator trong file Scope", ("indicator file scope", "indicator scope", "file scope", "scope"), 1),
    FieldSpec("indicator", "Chỉ tiêu", ("indicator", "chi tieu", "kpi", "metric", "measure", "ten chi tieu"), 4),
    FieldSpec("indicator_new", "Chỉ tiêu mới đề xuất", ("indicator new", "indicator proposed", "chi tieu moi", "chi tieu moi de xuat", "new indicator", "proposed indicator"), 2),
    FieldSpec("definition", "Định nghĩa/Mô tả", ("definition", "dinh nghia", "description", "mo ta", "meaning", "business definition", "glossary definition"), 4),
    FieldSpec("unit", "Đơn vị tính", ("unit", "dvt", "don vi tinh", "uom", "unit of measure"), 1),
    FieldSpec("application", "Application", ("application", "app", "he thong", "system", "source system"), 1),
    FieldSpec("api_file", "Loại API/File", ("api/file", "api file", "loai api", "loai api hay file", "api", "file type", "source type"), 4),
    FieldSpec("source_name", "Tên nguồn", ("source name", "ten nguon", "name source", "source", "data source"), 5),
    FieldSpec("raw_table", "Raw table", ("raw table", "raw zone table", "raw"), 1),
    FieldSpec("silver_table", "Silver table", ("silver table", "silver zone table", "silver"), 1),
    FieldSpec("gold_table", "Gold table", ("gold table", "gold zone table", "gold", "datamart table", "data mart table", "datamart"), 1),
    FieldSpec("calculation", "Công thức tính toán", ("calculation", "formula", "cong thuc", "cong thuc tinh toan", "logic", "business rule"), 2),
    FieldSpec("note", "Ghi chú", ("note", "notes", "ghi chu", "remark", "remarks", "comment"), 1),
    FieldSpec("is_new", "Mới?", ("moi", "is new", "new", "1 moi", "status new"), 1),
    FieldSpec("bsc_kpi", "BSC KPI?", ("bsc kpi", "bsc", "balanced scorecard"), 1),
    FieldSpec("clarify_source", "Làm rõ nguồn", ("lam ro nguon", "clarify source", "source clarification"), 1),
    FieldSpec("clarify_formula", "Làm rõ công thức", ("lam ro cong thuc", "clarify formula", "formula clarification"), 1),
    FieldSpec("priority", "Priority", ("priority", "muc uu tien", "do uu tien"), 1),
    FieldSpec("time_grain", "Thời gian", ("thoi gian", "time", "time grain", "period", "frequency", "grain"), 1),
    FieldSpec("dntv", "DNTV", ("dntv", "don vi thanh vien", "member company"), 1),
    FieldSpec("port_region", "Khu vực cảng", ("khu vuc cang", "port region", "region", "area"), 1),
    FieldSpec("synonyms", "Synonyms", ("synonyms", "synonym", "tu dong nghia", "alias", "aliases"), 1),
)

FIELD_BY_KEY: Dict[str, FieldSpec] = {field.key: field for field in FIELD_SPECS}
DESCRIPTION_FIELD_ORDER: Tuple[str, ...] = tuple(field.key for field in FIELD_SPECS)


# -----------------------------
# Text helpers
# -----------------------------

def clean(value: Any) -> str:
    """Convert Excel/API values to clean strings."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def strip_accents(text: str) -> str:
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return text.replace("đ", "d").replace("Đ", "D")


def normalize_text(text: Any) -> str:
    text = strip_accents(clean(text)).lower()
    text = text.replace("_", " ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def is_meaningful(value: Any) -> bool:
    text = normalize_text(value)
    raw = clean(value).strip()
    if text in INVALID_VALUES:
        return False
    if raw.startswith("#"):
        return False
    return bool(text)


def safe_name(text: Any, fallback: str = "Term", max_len: int = 120) -> str:
    """Make a safe OpenMetadata entity name."""
    raw = clean(text)
    if not raw:
        raw = fallback
    text = strip_accents(raw)
    text = re.sub(r"[^A-Za-z0-9_\-]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    if not text:
        text = fallback
    if re.match(r"^[0-9]", text):
        text = "T_" + text
    return text[:max_len]


def fqn_join(*parts: str) -> str:
    return ".".join(clean(part) for part in parts if clean(part))


def split_synonyms(text: str) -> List[str]:
    if not is_meaningful(text):
        return []
    parts = re.split(r"[,;\n|]+", text)
    result: List[str] = []
    seen: Set[str] = set()
    for part in parts:
        value = clean(part)
        key = normalize_text(value)
        if value and key not in seen:
            result.append(value)
            seen.add(key)
    return result


# -----------------------------
# Excel structure detection
# -----------------------------

def field_match_score(header_text: str, field: FieldSpec) -> int:
    """Return match score between one combined header text and a FieldSpec."""
    header = normalize_text(header_text)
    if not header:
        return 0

    best = 0
    for alias in field.aliases:
        alias_norm = normalize_text(alias)
        if not alias_norm:
            continue
        if header == alias_norm:
            best = max(best, 100 + field.weight + min(len(alias_norm), 30))
        elif re.search(rf"(^|\s){re.escape(alias_norm)}($|\s)", header):
            best = max(best, 80 + field.weight + min(len(alias_norm), 30))
        elif alias_norm in header:
            best = max(best, 60 + field.weight + min(len(alias_norm), 30))
    return best


def best_field_for_header(header_text: str, used_fields: Optional[Set[str]] = None) -> Optional[str]:
    used_fields = used_fields or set()
    scored: List[Tuple[int, str]] = []
    for field in FIELD_SPECS:
        if field.key in used_fields:
            continue
        score = field_match_score(header_text, field)
        if score > 0:
            scored.append((score, field.key))
    if not scored:
        return None
    scored.sort(reverse=True)
    return scored[0][1]


def row_alias_score(ws, row_idx: int, max_col: int) -> int:
    """Score a possible header row.

    The current row is weighted more strongly than the next row. This prevents
    merged/title rows above the real header from being selected only because
    the real header is directly underneath.
    """
    total = 0
    current_matched_fields: Set[str] = set()
    combined_matched_fields: Set[str] = set()

    for col_idx in range(1, max_col + 1):
        current = clean(ws.cell(row=row_idx, column=col_idx).value)
        below = clean(ws.cell(row=row_idx + 1, column=col_idx).value) if row_idx < (ws.max_row or 0) else ""
        combined = " / ".join(part for part in (current, below) if part)

        current_best = 0
        current_key = None
        combined_best = 0
        combined_key = None
        for field in FIELD_SPECS:
            score_current = field_match_score(current, field)
            if score_current > current_best:
                current_best = score_current
                current_key = field.key
            score_combined = field_match_score(combined, field)
            if score_combined > combined_best:
                combined_best = score_combined
                combined_key = field.key

        if current_best:
            total += current_best * 3
            if current_key:
                current_matched_fields.add(current_key)
        if combined_best:
            total += combined_best
            if combined_key:
                combined_matched_fields.add(combined_key)

    for critical in ("code", "term_name", "group", "indicator", "definition"):
        if critical in current_matched_fields:
            total += 120
        elif critical in combined_matched_fields:
            total += 30
    return total


def count_alias_matches_in_row(ws, row_idx: int, max_col: int) -> int:
    count = 0
    if row_idx < 1 or row_idx > (ws.max_row or 0):
        return 0
    for col_idx in range(1, max_col + 1):
        value = clean(ws.cell(row=row_idx, column=col_idx).value)
        if best_field_for_header(value):
            count += 1
    return count


def detect_header(ws, forced_header_row: Optional[int], forced_data_start_row: Optional[int]) -> Tuple[int, int, int]:
    """Return (header_start_row, header_end_row, data_start_row)."""
    max_col = min(ws.max_column or 1, DEFAULT_MAX_COLS)

    if forced_header_row:
        header_start = forced_header_row
    else:
        scan_to = min(ws.max_row or 1, DEFAULT_HEADER_SCAN_ROWS)
        best_row = 1
        best_score = -1
        for row_idx in range(1, scan_to + 1):
            score = row_alias_score(ws, row_idx, max_col)
            if score > best_score:
                best_score = score
                best_row = row_idx
        header_start = best_row

    # If the next row also looks like a header, treat it as a second header row.
    next_row_matches = count_alias_matches_in_row(ws, header_start + 1, max_col)
    current_row_matches = count_alias_matches_in_row(ws, header_start, max_col)
    header_end = header_start + 1 if next_row_matches >= 2 and next_row_matches >= current_row_matches * 0.4 else header_start

    data_start = forced_data_start_row or (header_end + 1)
    return header_start, header_end, data_start


def header_label_for_column(ws, col_idx: int, header_start: int, header_end: int) -> str:
    context_start = max(1, header_start - 1)
    parts: List[str] = []
    seen: Set[str] = set()
    for row_idx in range(context_start, header_end + 1):
        value = clean(ws.cell(row=row_idx, column=col_idx).value)
        key = normalize_text(value)
        if value and key not in seen:
            parts.append(value)
            seen.add(key)
    return " / ".join(parts) if parts else get_column_letter(col_idx)


def build_column_map(ws, header_start: int, header_end: int) -> Tuple[Dict[str, int], Dict[int, str]]:
    """Map logical field keys to Excel column indexes and keep readable labels."""
    max_col = min(ws.max_column or 1, DEFAULT_MAX_COLS)
    col_labels: Dict[int, str] = {}
    field_to_col: Dict[str, int] = {}
    used_fields: Set[str] = set()

    for col_idx in range(1, max_col + 1):
        label = header_label_for_column(ws, col_idx, header_start, header_end)
        col_labels[col_idx] = label
        matched = best_field_for_header(label, used_fields)
        if matched:
            field_to_col[matched] = col_idx
            used_fields.add(matched)

    return field_to_col, col_labels


def detect_best_sheet(workbook, requested_sheet: str, all_sheets: bool) -> List[str]:
    """Select sheet(s) to import."""
    visible_sheets = [ws.title for ws in workbook.worksheets]
    if all_sheets:
        return visible_sheets

    if requested_sheet in workbook.sheetnames:
        return [requested_sheet]

    # Keep the old default convenient: if Logic Tổng hợp does not exist in another
    # workbook, automatically select the sheet with the strongest glossary header.
    if requested_sheet != DEFAULT_SHEET:
        raise ValueError(f"Không tìm thấy sheet '{requested_sheet}'. Sheets hiện có: {visible_sheets}")

    best_title = visible_sheets[0]
    best_score = -1
    for ws in workbook.worksheets:
        max_col = min(ws.max_column or 1, DEFAULT_MAX_COLS)
        score = 0
        for row_idx in range(1, min(ws.max_row or 1, DEFAULT_HEADER_SCAN_ROWS) + 1):
            score = max(score, row_alias_score(ws, row_idx, max_col))
        if score > best_score:
            best_score = score
            best_title = ws.title
    return [best_title]


def get_cell(row_values: Sequence[Any], col_idx: Optional[int]) -> str:
    if not col_idx or col_idx < 1 or col_idx > len(row_values):
        return ""
    return clean(row_values[col_idx - 1])


def mapped_value(row_values: Sequence[Any], field_to_col: Dict[str, int], key: str) -> str:
    return get_cell(row_values, field_to_col.get(key))


def row_has_term(row_values: Sequence[Any], field_to_col: Dict[str, int]) -> bool:
    # Import a row when it has at least one business term signal.
    candidate_keys = ("code", "term_name", "display_name", "indicator", "indicator_new", "definition")
    for key in candidate_keys:
        if is_meaningful(mapped_value(row_values, field_to_col, key)):
            return True

    # Fallback for very simple spreadsheets with unrecognized headers: import if
    # at least 2 cells contain meaningful values.
    meaningful_cells = sum(1 for value in row_values if is_meaningful(value))
    return meaningful_cells >= 2


def build_description(
    source_file: str,
    sheet_name: str,
    excel_row: int,
    import_order: int,
    row_values: Sequence[Any],
    field_to_col: Dict[str, int],
    col_labels: Dict[int, str],
) -> str:
    """Build Markdown description for a glossary term."""
    lines = [
        f"**Nguồn file:** {source_file}",
        f"**Sheet:** {sheet_name}",
        f"**Dòng Excel:** {excel_row}",
        f"**Thứ tự import:** {import_order}",
    ]

    used_cols: Set[int] = set()
    for key in DESCRIPTION_FIELD_ORDER:
        col_idx = field_to_col.get(key)
        if not col_idx:
            continue
        value = get_cell(row_values, col_idx)
        if is_meaningful(value):
            label = FIELD_BY_KEY.get(key, FieldSpec(key, key, tuple())).label
            lines.append(f"**{label}:** {value}")
            used_cols.add(col_idx)

    extra_lines: List[str] = []
    for col_idx, raw_value in enumerate(row_values, start=1):
        if col_idx in used_cols:
            continue
        value = clean(raw_value)
        if is_meaningful(value):
            label = col_labels.get(col_idx) or get_column_letter(col_idx)
            extra_lines.append(f"- **{label}:** {value}")

    if extra_lines:
        lines.append("\n**Thông tin bổ sung từ các cột khác:**\n" + "\n".join(extra_lines))

    return "\n\n".join(lines)


# -----------------------------
# OpenMetadata API client
# -----------------------------

class OpenMetadataClient:
    def __init__(self, host: str, token: Optional[str], timeout: int = 45, retries: int = 2):
        self.base_url = self.normalize_api_base(host)
        self.timeout = timeout
        self.retries = max(0, retries)
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
        if token:
            self.session.headers.update({"Authorization": f"Bearer {token}"})

    @staticmethod
    def normalize_api_base(host: str) -> str:
        if not host:
            raise ValueError("OPENMETADATA_HOST is empty. Example: http://192.168.74.12:30085")
        host = host.strip().rstrip("/")
        if host.endswith("/api/v1"):
            return host
        if host.endswith("/api"):
            return host + "/v1"
        return host + "/api/v1"

    def request(self, method: str, path: str, json_body: Optional[dict] = None) -> dict:
        url = f"{self.base_url}{path}"
        last_response: Optional[requests.Response] = None
        for attempt in range(self.retries + 1):
            response = self.session.request(method, url, json=json_body, timeout=self.timeout)
            last_response = response
            if response.status_code < 500 and response.status_code != 429:
                break
            if attempt < self.retries:
                time.sleep(1.5 * (attempt + 1))

        assert last_response is not None
        response = last_response
        if response.status_code >= 400:
            hint = ""
            if response.status_code == 401:
                hint = "\nGợi ý: Token thiếu/sai. Kiểm tra OM_JWT_TOKEN."
            elif response.status_code == 403:
                hint = "\nGợi ý: Token không đủ quyền tạo/sửa Glossary."
            elif response.status_code == 400:
                hint = "\nGợi ý: Payload có thể chưa khớp version OpenMetadata hoặc tên term bị trùng/sai parent. Xem response body."
            elif response.status_code == 404:
                hint = "\nGợi ý: Kiểm tra URL OpenMetadata và endpoint /api/v1."
            raise RuntimeError(
                f"\nAPI ERROR {response.status_code} {method} {url}\n"
                f"Request body: {json_body}\n"
                f"Response body: {response.text}{hint}\n"
            )
        if not response.text.strip():
            return {}
        return response.json()

    def get(self, path: str) -> dict:
        return self.request("GET", path)

    def put(self, path: str, body: dict) -> dict:
        return self.request("PUT", path, body)

    def health_check(self) -> None:
        try:
            version = self.get("/system/version")
            print(f"✅ Connected to OpenMetadata: {self.base_url}")
            if version:
                print(f"   Version: {version}")
        except Exception as exc:
            print("⚠️ Không gọi được /system/version. Script vẫn sẽ thử import tiếp.")
            print(f"   Chi tiết: {exc}")

    def upsert_glossary(self, name: str, display_name: str, description: str, dry_run: bool = False) -> None:
        payload = {
            "name": name,
            "displayName": display_name,
            "description": description,
            "mutuallyExclusive": False,
        }
        if dry_run:
            print(f"[DRY-RUN] Upsert glossary: {name}")
            return
        self.put("/glossaries", payload)
        print(f"✅ Upsert glossary: {name}")

    def upsert_term(
        self,
        glossary: str,
        name: str,
        display_name: str,
        description: str,
        parent_fqn: Optional[str] = None,
        synonyms: Optional[List[str]] = None,
        dry_run: bool = False,
    ) -> None:
        payload: Dict[str, Any] = {
            "name": name,
            "glossary": glossary,
            "displayName": display_name or name,
            "description": description or f"Imported glossary term: {display_name or name}",
        }
        if parent_fqn:
            payload["parent"] = parent_fqn
        clean_synonyms = [s for s in (synonyms or []) if is_meaningful(s)]
        if clean_synonyms:
            payload["synonyms"] = clean_synonyms

        if dry_run:
            parent_note = f" under {parent_fqn}" if parent_fqn else ""
            print(f"[DRY-RUN] Upsert term: {glossary}.{name}{parent_note}")
            return

        self.put("/glossaryTerms", payload)
        parent_note = f" under {parent_fqn}" if parent_fqn else ""
        print(f"✅ Upsert term: {glossary}.{name}{parent_note}")


# -----------------------------
# Import logic
# -----------------------------

def make_ordered_name(prefix: str, order: int, raw_name: str, fallback: str, name_mode: str) -> str:
    base = safe_name(raw_name, fallback=fallback, max_len=100)
    if name_mode == "original":
        return base
    return safe_name(f"{prefix}{order:05d}_{base}", fallback=f"{prefix}{order:05d}", max_len=120)


def make_unique(name: str, used: Set[str], fallback_prefix: str = "Term") -> str:
    if name not in used:
        used.add(name)
        return name
    base = name[:105]
    counter = 2
    while True:
        candidate = safe_name(f"{base}_DUP{counter}", fallback=f"{fallback_prefix}_{counter}", max_len=120)
        if candidate not in used:
            used.add(candidate)
            return candidate
        counter += 1


def collect_sheet_rows(
    ws,
    header_row: Optional[int],
    data_start_row: Optional[int],
) -> Tuple[int, int, int, Dict[str, int], Dict[int, str], List[Tuple[int, List[Any]]]]:
    header_start, header_end, data_start = detect_header(ws, header_row, data_start_row)
    field_to_col, col_labels = build_column_map(ws, header_start, header_end)

    max_col = min(ws.max_column or 1, DEFAULT_MAX_COLS)
    rows: List[Tuple[int, List[Any]]] = []
    for excel_row in range(data_start, (ws.max_row or 0) + 1):
        row_values = [ws.cell(row=excel_row, column=col_idx).value for col_idx in range(1, max_col + 1)]
        if not row_has_term(row_values, field_to_col):
            continue
        rows.append((excel_row, row_values))

    return header_start, header_end, data_start, field_to_col, col_labels, rows


def import_excel_sheet(
    client: OpenMetadataClient,
    excel_path: Path,
    sheet_name: str,
    ws,
    glossary_name: str,
    dry_run: bool,
    sleep_seconds: float,
    header_row: Optional[int],
    data_start_row: Optional[int],
    name_mode: str,
    ungrouped_label: str,
) -> Tuple[int, int]:
    (
        header_start,
        header_end,
        data_start,
        field_to_col,
        col_labels,
        rows,
    ) = collect_sheet_rows(ws, header_row, data_start_row)

    source_base_name = excel_path.stem if sheet_name == DEFAULT_SHEET else f"{excel_path.stem}_{sheet_name}"
    source_term_name = safe_name(source_base_name, fallback="SourceFile")
    source_display = source_base_name
    source_fqn = fqn_join(glossary_name, source_term_name)

    print(f"\n📄 File: {excel_path.name} | Sheet: {sheet_name}")
    print(f"   Header rows: {header_start}-{header_end}; data starts at row {data_start}")
    print("   Column mapping:")
    for key, col_idx in sorted(field_to_col.items(), key=lambda item: item[1]):
        print(f"   - {get_column_letter(col_idx)} -> {key} ({FIELD_BY_KEY[key].label})")

    if not rows:
        print("⚠️ Không tìm thấy dòng dữ liệu hợp lệ để import.")
        return 0, 0

    client.upsert_term(
        glossary=glossary_name,
        name=source_term_name,
        display_name=source_display,
        description=f"Các thuật ngữ được import từ file Excel: {excel_path.name}, sheet: {sheet_name}.",
        parent_fqn=None,
        dry_run=dry_run,
    )

    group_name_by_value: Dict[str, str] = {}
    group_order_by_value: Dict[str, int] = {}
    used_group_names: Set[str] = set()
    used_term_names_by_group: Dict[str, Set[str]] = defaultdict(set)
    group_count = 0
    term_count = 0

    for import_order, (excel_row, row_values) in enumerate(rows, start=1):
        code = mapped_value(row_values, field_to_col, "code")
        term_name_value = mapped_value(row_values, field_to_col, "term_name")
        indicator = mapped_value(row_values, field_to_col, "indicator")
        indicator_new = mapped_value(row_values, field_to_col, "indicator_new")
        display_name_value = mapped_value(row_values, field_to_col, "display_name")
        definition = mapped_value(row_values, field_to_col, "definition")
        group_value = mapped_value(row_values, field_to_col, "group") or ungrouped_label

        group_key = normalize_text(group_value) or normalize_text(ungrouped_label)
        if group_key not in group_name_by_value:
            group_count += 1
            group_order_by_value[group_key] = group_count
            if name_mode == "original":
                raw_group_name = safe_name(group_value, fallback="Khac")
            else:
                raw_group_name = make_ordered_name("G", group_count, group_value, "Group", name_mode)
            group_name = make_unique(raw_group_name, used_group_names, fallback_prefix="Group")
            group_name_by_value[group_key] = group_name
            group_fqn = fqn_join(glossary_name, source_term_name, group_name)

            client.upsert_term(
                glossary=glossary_name,
                name=group_name,
                display_name=group_value,
                description=f"Nhóm/parent term '{group_value}' trong file {excel_path.name}, sheet {sheet_name}. Thứ tự nhóm: {group_count}.",
                parent_fqn=source_fqn,
                dry_run=dry_run,
            )
            if sleep_seconds:
                time.sleep(sleep_seconds)

        group_name = group_name_by_value[group_key]
        group_fqn = fqn_join(glossary_name, source_term_name, group_name)
        used_term_names = used_term_names_by_group[group_name]

        raw_term_name = code or term_name_value or display_name_value or indicator_new or indicator or definition or f"row_{excel_row}"
        term_name = make_ordered_name("T", import_order, raw_term_name, f"row_{excel_row}", name_mode)
        term_name = make_unique(term_name, used_term_names, fallback_prefix="Term")

        display_name = display_name_value or indicator_new or indicator or term_name_value or code or term_name
        description = build_description(
            source_file=excel_path.name,
            sheet_name=sheet_name,
            excel_row=excel_row,
            import_order=import_order,
            row_values=row_values,
            field_to_col=field_to_col,
            col_labels=col_labels,
        )

        synonyms = split_synonyms(mapped_value(row_values, field_to_col, "synonyms"))
        if indicator and display_name and normalize_text(indicator) != normalize_text(display_name):
            synonyms.append(indicator)
        if term_name_value and display_name and normalize_text(term_name_value) != normalize_text(display_name):
            synonyms.append(term_name_value)
        # Remove duplicate synonyms after adding automatic values.
        synonyms = split_synonyms("\n".join(synonyms))

        client.upsert_term(
            glossary=glossary_name,
            name=term_name,
            display_name=display_name,
            description=description,
            parent_fqn=group_fqn,
            synonyms=synonyms,
            dry_run=dry_run,
        )
        term_count += 1
        if sleep_seconds:
            time.sleep(sleep_seconds)

    print(f"Sheet '{sheet_name}': {group_count} groups, {term_count} terms")
    return group_count, term_count


def import_excel_file(
    client: OpenMetadataClient,
    excel_path: Path,
    glossary_name: str,
    requested_sheet: str,
    all_sheets: bool,
    dry_run: bool = False,
    sleep_seconds: float = 0.0,
    header_row: Optional[int] = None,
    data_start_row: Optional[int] = None,
    name_mode: str = "ordered",
    ungrouped_label: str = "Khác",
) -> Tuple[int, int]:
    workbook = load_workbook(excel_path, data_only=True, read_only=False)
    try:
        sheet_names = detect_best_sheet(workbook, requested_sheet, all_sheets)
        total_groups = 0
        total_terms = 0
        for sheet_name in sheet_names:
            ws = workbook[sheet_name]
            groups, terms = import_excel_sheet(
                client=client,
                excel_path=excel_path,
                sheet_name=sheet_name,
                ws=ws,
                glossary_name=glossary_name,
                dry_run=dry_run,
                sleep_seconds=sleep_seconds,
                header_row=header_row,
                data_start_row=data_start_row,
                name_mode=name_mode,
                ungrouped_label=ungrouped_label,
            )
            total_groups += groups
            total_terms += terms
        return total_groups, total_terms
    finally:
        workbook.close()


def resolve_excel_paths(paths: Iterable[str]) -> List[Path]:
    resolved: List[Path] = []
    for p in paths:
        path = Path(p)
        if not path.exists():
            raise FileNotFoundError(f"Không tìm thấy file: {p}")
        if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            raise ValueError(f"File không phải Excel được hỗ trợ: {p}")
        resolved.append(path)
    return resolved


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Import VIMC/Business Glossary Excel files into OpenMetadata Glossary."
    )
    parser.add_argument(
        "--excel",
        action="append",
        help="Excel file path. Có thể dùng nhiều lần: --excel file1.xlsx --excel file2.xlsx",
    )
    parser.add_argument(
        "--glossary",
        default=DEFAULT_GLOSSARY,
        help=f"OpenMetadata Glossary name. Default: {DEFAULT_GLOSSARY}",
    )
    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET,
        help=(
            f"Tên sheet cần đọc. Default: {DEFAULT_SHEET}. Nếu sheet mặc định không tồn tại, "
            "script tự chọn sheet có header giống glossary nhất."
        ),
    )
    parser.add_argument(
        "--all-sheets",
        action="store_true",
        help="Import tất cả sheets trong mỗi workbook thay vì chỉ 1 sheet.",
    )
    parser.add_argument(
        "--host",
        default=os.getenv("OPENMETADATA_HOST"),
        help="OpenMetadata host, e.g. http://192.168.74.12:30085",
    )
    parser.add_argument(
        "--token",
        default=os.getenv("OM_JWT_TOKEN"),
        help="OpenMetadata JWT token / Personal Access Token.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Chỉ đọc file và in kế hoạch import, không gọi API OpenMetadata.",
    )
    parser.add_argument(
        "--sleep",
        type=float,
        default=0.0,
        help="Nghỉ giữa mỗi API call, ví dụ 0.05 nếu server yếu.",
    )
    parser.add_argument(
        "--header-row",
        type=int,
        default=None,
        help="Ép dòng header nếu file đặc biệt. Bỏ trống để tự dò.",
    )
    parser.add_argument(
        "--data-start-row",
        type=int,
        default=None,
        help="Ép dòng bắt đầu dữ liệu nếu file đặc biệt. Bỏ trống để tự dò sau header.",
    )
    parser.add_argument(
        "--name-mode",
        choices=("ordered", "original"),
        default="ordered",
        help=(
            "ordered: thêm prefix G/T để OpenMetadata sort đúng thứ tự Excel; "
            "original: dùng nguyên Code/Name làm entity name. Default: ordered."
        ),
    )
    parser.add_argument(
        "--ungrouped-label",
        default="Khác",
        help="Tên group mặc định khi dòng dữ liệu không có Group/Parent. Default: Khác.",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=45,
        help="Timeout mỗi API call, đơn vị giây. Default: 45.",
    )
    parser.add_argument(
        "--retries",
        type=int,
        default=2,
        help="Số lần retry cho lỗi 429/5xx. Default: 2.",
    )
    args = parser.parse_args()

    excel_inputs = args.excel or DEFAULT_FILES

    try:
        excel_paths = resolve_excel_paths(excel_inputs)
    except Exception as exc:
        print(f"❌ {exc}")
        print("Gợi ý: đặt script cùng thư mục với file Excel, hoặc truyền đường dẫn đầy đủ bằng --excel.")
        return 1

    glossary_name = safe_name(args.glossary, fallback=DEFAULT_GLOSSARY)

    print("\n=== VIMC/OpenMetadata Glossary Import ===")
    print(f"Glossary: {glossary_name}")
    print(f"Requested sheet: {args.sheet}")
    print(f"Name mode: {args.name_mode}")
    print("Files:")
    for p in excel_paths:
        print(f"- {p}")

    if not args.host and not args.dry_run:
        print("\n❌ Thiếu OPENMETADATA_HOST. Ví dụ PowerShell:")
        print("$env:OPENMETADATA_HOST='http://192.168.74.12:30085'")
        return 1

    if not args.token and not args.dry_run:
        print("\n⚠️ OM_JWT_TOKEN đang trống. Nếu server yêu cầu đăng nhập, API có thể lỗi 401.")
        print("PowerShell ví dụ:")
        print("$env:OM_JWT_TOKEN='PASTE_TOKEN_CUA_BAN'")

    client = OpenMetadataClient(
        host=args.host or "http://localhost:8585",
        token=args.token,
        timeout=args.timeout,
        retries=args.retries,
    )
    if not args.dry_run:
        client.health_check()

    client.upsert_glossary(
        name=glossary_name,
        display_name=args.glossary,
        description="Business Glossary imported automatically from Excel files using Python.",
        dry_run=args.dry_run,
    )

    total_groups = 0
    total_terms = 0
    for p in excel_paths:
        try:
            groups, terms = import_excel_file(
                client=client,
                excel_path=p,
                glossary_name=glossary_name,
                requested_sheet=args.sheet,
                all_sheets=args.all_sheets,
                dry_run=args.dry_run,
                sleep_seconds=args.sleep,
                header_row=args.header_row,
                data_start_row=args.data_start_row,
                name_mode=args.name_mode,
                ungrouped_label=args.ungrouped_label,
            )
        except Exception as exc:
            print(f"\n❌ Lỗi khi import file '{p}': {exc}")
            return 1
        total_groups += groups
        total_terms += terms

    print(f"Tổng group terms: {total_groups}")
    print(f"Tổng business/indicator terms: {total_terms}")
    if not args.dry_run:
        print(f"Kiểm tra trên OpenMetadata UI: Govern → Glossary → {glossary_name}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
